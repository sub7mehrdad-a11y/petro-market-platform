"""
استخراج لیست شرکت‌کنندگان یک نمایشگاه از صفحه‌ی «Exhibitor List» آن و فیلترشان
به یک فهرست کوتاه و قابل‌استفاده از خریداران بالقوه‌ی گریدهای مختلف جوش شیرین.

چرا مرورگر لازم نیست: این‌جور صفحات (حداقل الگوی دیده‌شده در Gulfood) لیست را
با یک درخواست AJAX داخلی (fetchExhibitors) به‌صورت قطعه‌های HTML می‌گیرند، نه
با رندر جاوااسکریپت سنگین. با httpx مستقیم همان درخواست را می‌زنیم — سریع‌تر و
بدون نیاز به مرورگر داخلی (که طبق تجربه‌ی قبلی این پروژه باعث کرش سیستم کاربر
شده؛ به همین خاطر این‌جا اصلاً به آن دست نمی‌زنیم).

چرا دو مرحله (قانون + LLM)، نه یکی:
  - یک نمایشگاه غذایی مثل Gulfood تقریباً همه‌ی شرکت‌کننده‌هاش برچسب‌های عمومی
    غذایی دارن (Bakery/Beverages/Dairy روی اکثر پروفایل‌ها هست). فیلتر کردن با
    همون برچسب‌های عمومی عملاً هیچی رو کنار نمی‌ذاره (در تست اول روی Gulfood
    2026: از ۶۵۰۰ نمایشگاه‌دهنده، ۵۱۶۱ تا رد شدن).
  - راه‌حل: مرحله‌ی اول فقط با زیردسته‌های خیلی مشخص (نه برچسب‌های کلی) یک
    زیرمجموعه‌ی کوچک‌تر می‌سازه (فیلتر ارزون و بدون AI). مرحله‌ی دوم همون
    زیرمجموعه‌ی کوچیک رو با یک مدل زبانی (Groq، طبق قرارداد بقیه‌ی ربات‌های این
    پروژه) واقعاً می‌خونه و قضاوت می‌کنه — چیزی که فیلتر کلیدواژه‌ای اصلاً
    نمی‌تونه انجام بده (تشخیص این‌که این شرکت مشتری/شریک تجاری بالقوه‌ی جوش
    شیرینه یا نه، و کدوم گرید به کارش می‌خوره).

نکته‌ی مهم درباره‌ی بازرگان‌ها: برخلاف حدس اولیه، شرکت‌های بازرگانی/توزیع‌کننده
از فیلتر نهایی حذف نمی‌شن — طبق نظر کاربر (صاحب کسب‌وکار)، بازرگان‌ها اتفاقاً
مشتری‌های بهتری هستن چون خودشون توزیع‌کننده‌ی بازار محلی‌شونن و می‌تونن جوش
شیرین رو به‌عنوان یک قلم جدید به سبد کالای خودشون اضافه کنن. پس هم تولیدکننده
(مصرف‌کننده‌ی مستقیم به‌عنوان ماده‌ی اولیه) و هم بازرگان/توزیع‌کننده‌ی حوزه‌ی
غذایی، هر دو کاندید مرتبط محسوب می‌شن؛ فقط چیزهای کاملاً بی‌ربط (تجهیزات،
فناوری، خدمات غیرغذایی) کنار گذاشته می‌شن.

مراحل خط لوله:
  ۱. گرفتن صفحه‌ی اصلی نمایشگاه و استخراج درخت دسته‌بندی محصولات.
  ۲. انتخاب فقط زیردسته‌های (leaf) خیلی مشخصی که با مصارف شناخته‌شده‌ی جوش
     شیرین در صنایع غذایی هم‌خوانی دارن (نه برچسب‌های سرشاخه‌ی کلی).
  ۳. صفحه‌بندی روی fetchExhibitors با همان فیلتر (سمت سرور) تا حجم دانلود کم بماند.
  ۴. روی نتیجه، یک پاس Groq برای تأیید نهایی ربط (تولیدکننده یا بازرگان/
     توزیع‌کننده‌ی حوزه‌ی غذایی، هر دو خوبه) + گرید پیشنهادی.
  ۵. خروجی JSON در data/exhibition_leads/<اسلاگ‌رویداد>.json ذخیره می‌شود.

محدودیت شناخته‌شده: این اسکریپت فقط الگوی همین پلتفرم (fetchExhibitors +
dropdownContent) را می‌فهمد. اگر نمایشگاه بعدی از یک CMS متفاوت استفاده کند،
تابع fetch_category_tree/parse_exhibitor_fragment باید برای آن سایت بازنویسی
شود — عمداً یک اسکریپر «همه‌کاره‌ی حدسی» برای ساختار ناشناخته نساختیم.

چک‌پوینت (مهم): مرحله‌ی ۴ (پاس LLM) روی چند هزار شرکت می‌تونه دقیقه‌ها طول
بکشه، و این محیط اجرا (نشست ابزار) چند بار وسط کار قطع شده و کل پیشرفت از بین
رفته. برای همین، بعد از هر دسته‌ی LLM_BATCH_SIZE‌تایی، نتیجه فوری در فایل
<اسلاگ‌رویداد>.progress.json ذخیره می‌شه و data/exhibition_leads/<اسلاگ>.json هم
همون لحظه بازنویسی می‌شه. اجرای بعدی همین دستور، شرکت‌های قبلاً پردازش‌شده رو
(چه تأییدشده چه ردشده) دوباره به مدل نمی‌فرسته — یعنی هر بار قطع بشه، دفعه‌ی
بعد از همون‌جا ادامه پیدا می‌کنه، نه از صفر.

چرا Gemini پیش‌فرضه نه Groq: توی اولین اجرای کامل روی ۲۵۳۲ کاندید، فقط با ۲۴۹
شرکت (کمتر از ۱۰٪ کار)، کل سهمیه‌ی روزانه‌ی Groq (۲۰۰٬۰۰۰ توکن/روز، مشترک با
کل حساب) تموم شد. این حساب رو بخش «پاسخ هوشمند» زنده‌ی سایت (web/app/api/ask)
هم استفاده می‌کنه، پس مصرف سنگین این‌جا می‌تونست اون قابلیت رو برای بازدیدکننده‌های
واقعی سایت امروز از کار بندازه. Gemini سهمیه‌ی جدا داره (همونی که بقیه‌ی
ربات‌های روزانه‌ی پروژه ازش استفاده می‌کنن)، پس برای این‌جور کار حجیم امن‌تره.
"""

import os
import re
import sys
import json
import time
import argparse
from datetime import datetime, timezone
from urllib.parse import urlparse

import httpx
import openpyxl
from openpyxl.styles import Alignment, Font
from bs4 import BeautifulSoup
from google import genai

sys.path.insert(0, os.path.dirname(__file__))
from env_utils import load_env  # noqa: E402
from groq_utils import groq_generate  # noqa: E402

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "exhibition_leads")

# پوشه‌ی محلی کاربر (بیرون از مخزن گیت، عمداً) — طبق درخواست کاربر، برای هر
# نمایشگاه یک فایل اکسل جدا این‌جا ساخته می‌شه تا مستقیم در ویندوز قابل‌مرور باشه.
# روی GitHub Actions (لینوکس، بدون این مسیر) با EXHIBITOR_XLSX_DIR به یک مسیر
# داخل خودِ مخزن override می‌شه تا قابل commit/دانلود به‌عنوان artifact باشه.
EXTERNAL_XLSX_DIR = os.environ.get("EXHIBITOR_XLSX_DIR") or r"E:\مهرداد\سپهران\نمایشگاه دارها"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "X-Requested-With": "XMLHttpRequest",
}

REQUEST_DELAY_SEC = 0.4
PAGE_LIMIT = 100

# فقط زیردسته‌های (leaf) خیلی مشخص که مصرف واقعی جوش شیرین در آن‌ها شناخته‌شده‌ست
# — عمداً از برچسب‌های سرشاخه‌ی کلی (مثل خودِ "Bakery" یا "Beverages") استفاده
# نکردیم چون تقریباً همه‌ی نمایشگاه‌دهنده‌های یک اکسپوی غذایی حداقل یکی از
# سرشاخه‌ها رو دارن. هر ورودی: (برچسب دقیق دسته روی سایت، گرید/کاربرد فارسی).
LEAF_CATEGORY_TO_GRADE = {
    "baking supplies": "گرید غذایی – پودر بکینگ / عامل ورآمدن خمیر",
    "baked goods": "گرید غذایی – نانوایی (نان و محصولات پخته)",
    "gluten-free baked goods": "گرید غذایی – نانوایی بدون گلوتن",
    "cakes / dessert": "گرید غذایی – کیک و دسر",
    "frozen baked goods": "گرید غذایی – محصولات پخته‌ی منجمد",
    "savoury baked products": "گرید غذایی – اسنک پخته‌ی شور",
    "cheese": "گرید غذایی – پنیر فرآوری‌شده (تنظیم اسیدیته)",
    "processed meat": "گرید غذایی – فرآوری گوشت (تنظیم pH/نرم‌کننده)",
    "carbonated soft drinks": "گرید غذایی – نوشابه‌سازی",
    "fortified & functional products": "گرید غذایی/دارویی – محصولات فراسودمند",
    "health & wellness products": "گرید غذایی/دارویی – سلامت و مکمل (جوشان/آنتی‌اسید)",
}

LLM_SYSTEM_PROMPT = """
تو یک تحلیل‌گر بازاریابی B2B برای یک تولیدکننده‌ی جوش شیرین (سدیم بی‌کربنات،
گرید غذایی) هستی که دنبال مشتری/شریک تجاری بالقوه در بین شرکت‌کننده‌های یک
نمایشگاه غذایی می‌گرده. برای هر شرکت زیر تشخیص بده آیا احتمالاً یکی از این دو
دسته است یا نه:
  (الف) تولیدکننده‌ای که در محصولاتش از جوش شیرین به‌عنوان ماده‌ی اولیه استفاده
        می‌کند (نانوایی، نوشابه‌سازی، لبنیات، فرآوری گوشت، مکمل/دارویی، ...).
  (ب) بازرگان/واردکننده/توزیع‌کننده‌ی محصولات یا مواد اولیه‌ی غذایی در کشور
        خودش — این‌ها هم کاندید خوبی هستن چون می‌تونن جوش شیرین رو به‌عنوان یک
        قلم جدید به سبد کالای توزیعی‌شون در بازار محلی اضافه کنن. بازرگان بودن
        به‌تنهایی دلیل رد نیست، حتی ترجیح داده می‌شه.

فقط شرکت‌هایی رو رد کن (relevant=false) که واقعاً هیچ ربطی به تجارت/تولید مواد
غذایی یا مواد اولیه‌ی غذایی ندارن (مثلاً فقط تجهیزات/ماشین‌آلات، فناوری، خدمات
حرفه‌ای، بسته‌بندی، یا رویدادها/تبلیغات می‌فروشن).

برای هر شرکت در آرایه‌ی ورودی، یک آبجکت با این ساختار دقیق در پاسخ بگنجان:
{
  "index": <همان index ورودی>,
  "relevant": true/false,
  "company_type": "manufacturer" یا "trader" یا null (اگر relevant=false),
  "grade": "گرید/کاربرد پیشنهادی جوش شیرین به فارسی (مثلاً 'گرید غذایی – نانوایی')، اگر relevant=false باشد null",
  "reason": "یک جمله‌ی کوتاه فارسی برای دلیل تصمیم"
}

فقط یک JSON array از این آبجکت‌ها برگردون، دقیقاً به تعداد ورودی‌ها، بدون
markdown fence و بدون توضیح اضافه.
"""

LLM_BATCH_SIZE = 15
GEMINI_MODEL = "gemini-3.6-flash"


def make_generate_fn(engine: str):
    """generate_fn(system_instruction, input_text) -> raw_text، مستقل از موتور زیرین."""
    if engine == "groq":
        # کلید اختصاصی این اسکریپت — تا سهمیه‌ی روزانه‌اش با ربات ترانزیت یا
        # بخش /ask زنده‌ی سایت قاطی نشه (۲۰۲۶-۰۹-۰۲: اولین اجرا با کلید مشترک
        # کل سهمیه‌ی روزانه‌ی حساب رو با کمتر از ۱۰٪ کار مصرف کرد).
        api_key = os.environ.get("GROQ_API_KEY_EXHIBITOR")
        if not api_key:
            raise SystemExit("GROQ_API_KEY_EXHIBITOR تنظیم نشده است.")
        return lambda sys_p, inp: groq_generate(
            system_instruction=sys_p, input=inp, reasoning_effort="low", max_tokens=2500,
            api_key=api_key,
        )

    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        raise SystemExit("GEMINI_API_KEY تنظیم نشده است.")
    client = genai.Client(api_key=api_key)

    def _generate(sys_p, inp):
        interaction = client.interactions.create(model=GEMINI_MODEL, system_instruction=sys_p, input=inp)
        return interaction.output_text

    return _generate


def parse_event_slug(list_url: str) -> tuple[str, str]:
    """از URL صفحه‌ی exhibitor list، base_url و اسلاگ رویداد را جدا می‌کند."""
    parsed = urlparse(list_url)
    base_url = f"{parsed.scheme}://{parsed.netloc}"
    parts = [p for p in parsed.path.split("/") if p]
    if not parts:
        raise ValueError(f"نتوانستم اسلاگ رویداد را از URL استخراج کنم: {list_url}")
    event_slug = parts[0]
    return base_url, event_slug


def fetch_category_tree(client: httpx.Client, base_url: str, event_slug: str, list_url: str):
    """درخت دسته‌بندی محصولات را از بخش dropdownContent صفحه‌ی اصلی می‌گیرد."""
    resp = client.get(list_url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    html = resp.text

    if "fetchExhibitors" not in html and "exhibitorlist.js" not in html:
        raise RuntimeError(
            "این صفحه ساختار شناخته‌شده (fetchExhibitors / exhibitorlist.js) را ندارد؛ "
            "این اسکریپت فقط این الگوی خاص پلتفرم نمایشگاهی را می‌فهمد و نیاز به تطبیق "
            "دستی برای این سایت دارد."
        )

    soup = BeautifulSoup(html, "html.parser")
    dropdown = soup.find(id="dropdownContent")
    categories = []  # [{id, label, parent_id}]
    if dropdown:
        for cat_input in dropdown.select("input.category"):
            label_tag = cat_input.find_parent("label")
            label_text = label_tag.get_text(strip=True) if label_tag else ""
            categories.append({"id": cat_input.get("value"), "label": label_text, "parent_id": None})
        for sub_input in dropdown.select("input.subcategory-box"):
            label_tag = sub_input.find_parent("label")
            label_text = label_tag.get_text(strip=True) if label_tag else ""
            categories.append({
                "id": sub_input.get("value"),
                "label": label_text,
                "parent_id": sub_input.get("data-parent"),
            })
    return categories, html


def select_relevant_categories(categories: list[dict]) -> tuple[list[str], list[str]]:
    """
    فقط زیردسته‌ها (leaf) را با LEAF_CATEGORY_TO_GRADE مقایسه می‌کند — برچسب‌های
    سرشاخه (parent_id=None) عمداً نادیده گرفته می‌شن چون خیلی کلی‌ن.
    """
    sub_ids = []
    for cat in categories:
        if cat["parent_id"] is None or not cat["label"]:
            continue
        if cat["label"].strip().lower() in LEAF_CATEGORY_TO_GRADE:
            sub_ids.append(cat["id"])
    return [], sub_ids


def build_fetch_payload(start: int, event_slug: str, new_category: str, new_sub_category: str) -> dict:
    return {
        "limit": PAGE_LIMIT, "start": start, "keyword_search": "", "cuntryId": "",
        "event_prod_cat_id": "", "exb_listed_as": "", "InitialKey": "",
        "selected_event_id": "", "start_up_exhibitors": "", "pav_country_id": "",
        "type": "", "vacancies": "", "product_search": "", "search_by_hall": "",
        "new_category": new_category, "new_sub_category": new_sub_category,
        "new_sub_sub_category": "", "search_by_venue": "", "event_sector_value": "",
    }


def parse_exhibitor_fragment(html: str, base_url: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    items = []
    for row in soup.select("div.item.list-group-item"):
        heading = row.select_one("h4.heading")
        if not heading:
            continue
        # اسم‌ها گاهی با یک پیشوند نویزی ("0 "، "0\x00 "، "1 ") شروع می‌شوند —
        # ظاهراً یک پرچم داخلی سایت (شاید نشان‌گر «نمایشگاه‌دهنده‌ی جدید») است
        # که در CSS مخفی می‌شود ولی در متن خام باقی می‌ماند.
        name = re.sub(r"^[\d\s\x00!]+(?=[A-Za-z\"'ÀÖ])", "", heading.get_text(" ", strip=True)).strip()
        if not name:
            name = heading.get_text(" ", strip=True).strip()

        stand = ""
        country = ""
        web_block = row.select_one("div.web")
        if web_block:
            ps = web_block.find_all("p", recursive=False)
            if ps:
                stand = ps[0].get_text(strip=True)
            span = web_block.select_one("span")
            if span:
                country = span.get_text(strip=True)

        desc_tag = row.select_one("p.list-group-item-text span")
        description = desc_tag.get_text(" ", strip=True) if desc_tag else ""

        categories = [li.get_text(strip=True) for li in row.select("ul.sector_block li")]

        profile_link = row.select_one("a.btn[href*='ExbDetails']")
        profile_url = profile_link["href"] if profile_link and profile_link.has_attr("href") else None
        if profile_url and profile_url.startswith("/"):
            profile_url = base_url + profile_url

        items.append({
            "name": name,
            "stand": stand,
            "country": country,
            "description": description,
            "categories": categories,
            "profile_url": profile_url,
        })
    return items


def fetch_all_exhibitors(client: httpx.Client, base_url: str, event_slug: str, list_url: str,
                          new_category: str, new_sub_category: str, max_items: int | None) -> list[dict]:
    fetch_url = f"{base_url}/{event_slug}/Exhibitor/fetchExhibitors"
    headers = {**HEADERS, "Referer": list_url}

    all_items, start = [], 0
    while True:
        payload = build_fetch_payload(start, event_slug, new_category, new_sub_category)
        resp = client.post(fetch_url, data=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        if not resp.text.strip():
            break

        batch = parse_exhibitor_fragment(resp.text, base_url)
        if not batch:
            break
        all_items.extend(batch)
        print(f"  [OK] start={start}: {len(batch)} مورد (جمع: {len(all_items)})")

        if max_items and len(all_items) >= max_items:
            all_items = all_items[:max_items]
            break

        start += PAGE_LIMIT
        time.sleep(REQUEST_DELAY_SEC)

    return all_items


def dedupe_by_profile(items: list[dict]) -> list[dict]:
    seen, out = set(), []
    for it in items:
        key = it.get("profile_url") or it["name"]
        if key in seen:
            continue
        seen.add(key)
        out.append(it)
    return out


def tag_rule_based_grade(items: list[dict]) -> list[dict]:
    """گرید اولیه را از روی اولین زیردسته‌ی منطبق در LEAF_CATEGORY_TO_GRADE می‌گذارد."""
    out = []
    for it in items:
        grade = None
        for cat in it.get("categories", []):
            grade = LEAF_CATEGORY_TO_GRADE.get(cat.strip().lower())
            if grade:
                break
        out.append({**it, "rule_based_grade": grade})
    return out


# ------------------------------------------------------------- استخراج تماس
#
# صفحه‌ی پروفایل نمایشگاه فقط لینک «وب‌سایت» شرکت رو داره (li.social_website)،
# نه ایمیل/تلفن مستقیم. برای اون‌ها باید خودِ سایت شرکت رو گرفت — که چون هر
# شرکت سایت متفاوتی داره (ساختار متفاوت، بعضی پشت Cloudflare/بلاک ربات، بعضی
# فقط فرم تماس دارن)، پوشش این مرحله ذاتاً جزئی و بهترین‌تلاش‌محوره، نه تضمینی.

EMAIL_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9._%+-]*@[A-Za-z0-9][A-Za-z0-9.-]*\.[A-Za-z]{2,}")
# پسوندهای فایل تصویری/فونت که گاهی به‌اشتباه شبیه ایمیل پارس می‌شن (مثل logo@2x.png)
EMAIL_JUNK_SUFFIX = re.compile(r"\.(png|jpe?g|gif|svg|webp|css|js|woff2?)$", re.IGNORECASE)
WHATSAPP_RE = re.compile(r"(?:https?://)?(?:api\.)?wa\.me/(\+?\d[\d\-]{6,15})", re.IGNORECASE)
WHATSAPP_LINK_RE = re.compile(r"(?:https?://)?(?:api\.)?whatsapp\.com/send\?phone=(\+?\d[\d\-]{6,15})", re.IGNORECASE)
TEL_RE = re.compile(r'href=["\']tel:(\+?[\d\-\s()]{6,20})["\']', re.IGNORECASE)

CONTACT_PATHS = ["", "/contact", "/contact-us", "/contactus", "/en/contact", "/en/contact-us"]


def fetch_exhibitor_website(client: httpx.Client, profile_url: str) -> str | None:
    """لینک «VISIT WEBSITE» را از صفحه‌ی پروفایل نمایشگاه استخراج می‌کند."""
    if not profile_url:
        return None
    try:
        resp = client.get(profile_url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
    except httpx.HTTPError:
        return None
    soup = BeautifulSoup(resp.text, "html.parser")
    link = soup.select_one("li.social_website a[href]")
    if not link:
        return None
    # بعضی صفحات یک گلیف آیکون‌فونت (کاراکتر private-use یونیکد) قبل از خودِ
    # لینک دارن که httpx.InvalidURL می‌ده — همون چیزی که یک‌بار کل اجرا رو
    # کرش داد؛ این‌جا حذفشون می‌کنیم.
    href = re.sub(r"[^\x00-\x7F]+", "", link["href"]).strip()
    return href or None


def extract_contacts_from_html(html: str) -> dict:
    emails = sorted({m for m in EMAIL_RE.findall(html) if not EMAIL_JUNK_SUFFIX.search(m)})
    whatsapp = sorted({m for m in WHATSAPP_RE.findall(html)} | {m for m in WHATSAPP_LINK_RE.findall(html)})
    phones = sorted(set(TEL_RE.findall(html)))
    return {"emails": emails[:3], "whatsapp": whatsapp[:2], "phones": phones[:3]}


def enrich_company_contacts(client: httpx.Client, profile_url: str) -> dict:
    """
    از پروفایل نمایشگاه لینک سایت رو می‌گیره، بعد خودِ سایت (و ۱-۲ مسیر رایج
    صفحه‌ی تماس) رو برای ایمیل/واتساپ/تلفن می‌گرده. نتیجه همیشه یک dict برمی‌گردونه
    (حتی اگه چیزی پیدا نشه) تا در progress قابل ثبت باشه و دوباره تلاش نشه.
    """
    result = {"website": None, "email": None, "whatsapp": None, "contact_note": None}

    website = fetch_exhibitor_website(client, profile_url)
    if not website:
        result["contact_note"] = "لینک وب‌سایت در پروفایل نمایشگاه پیدا نشد"
        return result
    result["website"] = website

    found = {"emails": [], "whatsapp": [], "phones": []}
    checked_any = False
    for path in CONTACT_PATHS:
        try:
            resp = client.get(website.rstrip("/") + path, headers=HEADERS, timeout=12, follow_redirects=True)
            checked_any = True
        except Exception:  # noqa: BLE001 — سایت‌های بیرونی غیرقابل‌اعتمادن؛ هیچ خطایی نباید کل اجرا را کرش بدهد
            continue
        if resp.status_code >= 400:
            continue
        c = extract_contacts_from_html(resp.text)
        found["emails"] = found["emails"] or c["emails"]
        found["whatsapp"] = found["whatsapp"] or c["whatsapp"]
        found["phones"] = found["phones"] or c["phones"]
        if found["emails"] and (found["whatsapp"] or found["phones"]):
            break  # هرچی لازم بود پیدا شد، سراغ مسیر بعدی نریم

    if found["emails"]:
        result["email"] = found["emails"][0]
    if found["whatsapp"]:
        result["whatsapp"] = found["whatsapp"][0]
    elif found["phones"]:
        result["whatsapp"] = found["phones"][0]  # تلفن معمولی، به‌عنوان جایگزین واتساپ

    if not checked_any:
        result["contact_note"] = "سایت شرکت در دسترس نبود (تایم‌اوت/خطا)"
    elif not result["email"] and not result["whatsapp"]:
        result["contact_note"] = "سایت باز شد ولی ایمیل/تماس مستقیم رویش پیدا نشد"

    return result


def enrich_contacts_pass(progress: dict, progress_path: str, on_batch_done, flush_every: int = 20,
                          deadline: float | None = None) -> None:
    """
    روی همه‌ی شرکت‌های relevant=true در progress که هنوز غنی‌سازی تماس نشدن
    (کلید contact_checked ندارن) اجرا می‌شه. هر flush_every تا، فوری ذخیره می‌کنه —
    مثل llm_refine، اگه وسط راه قطع بشه چیزی از دست نمی‌ره.
    """
    todo = [(k, v) for k, v in progress.items() if v.get("relevant") and not v.get("contact_checked")]
    if not todo:
        print("  همه‌ی شرکت‌های مرتبط قبلاً برای تماس بررسی شده‌اند.")
        return
    print(f"  {len(todo)} شرکت مرتبط برای استخراج ایمیل/وب‌سایت/واتساپ بررسی می‌شه.")

    with httpx.Client(follow_redirects=True) as client:
        for i, (key, record) in enumerate(todo, start=1):
            if deadline and time.time() > deadline:
                save_progress(progress_path, progress)
                on_batch_done()
                print(f"  [DEADLINE] زمان تمام شد؛ {len(todo) - i + 1} شرکت باقی‌مانده برای اجرای بعدی می‌مونه.")
                return

            try:
                contacts = enrich_company_contacts(client, record.get("profile_url"))
            except Exception as e:  # noqa: BLE001 — یک شرکت خراب نباید کل پاس تماس‌گیری را متوقف کند
                contacts = {"website": None, "email": None, "whatsapp": None,
                            "contact_note": f"خطای غیرمنتظره: {e}"}
            progress[key] = {**record, **contacts, "contact_checked": True}

            if i % flush_every == 0 or i == len(todo):
                save_progress(progress_path, progress)
                on_batch_done()
                print(f"  [OK] {i}/{len(todo)} شرکت بررسی شد")


def extract_json_array(text: str) -> list:
    fence_match = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.DOTALL)
    raw = fence_match.group(1) if fence_match else text
    array_match = re.search(r"\[.*\]", raw, re.DOTALL)
    if not array_match:
        raise ValueError("پاسخ مدل شامل JSON array نبود:\n" + text[:500])
    return json.loads(array_match.group(0))


def candidate_key(item: dict) -> str:
    return item.get("profile_url") or item["name"]


def atomic_write_json(path: str, data) -> None:
    """اول در یک فایل موقت می‌نویسه بعد جابه‌جا می‌کنه — اگه وسط نوشتن قطع بشیم،
    فایل قبلی سالم می‌مونه (نه یک JSON نصفه‌ونیمه‌ی خراب)."""
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)


def load_progress(progress_path: str) -> dict:
    if not os.path.exists(progress_path):
        return {}
    with open(progress_path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_progress(progress_path: str, progress: dict) -> None:
    atomic_write_json(progress_path, progress)


COMPANY_TYPE_FA = {"manufacturer": "تولیدکننده", "trader": "بازرگان/توزیع‌کننده"}


def safe_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "-", name).strip()


def write_xlsx(companies: list[dict], event_name: str) -> str:
    """یک فایل اکسل جدا برای این نمایشگاه در EXTERNAL_XLSX_DIR می‌سازه/به‌روز می‌کنه —
    طبق درخواست کاربر، مستقل از JSON داخل مخزن، برای مرور مستقیم در ویندوز."""
    os.makedirs(EXTERNAL_XLSX_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "شرکت‌ها"
    ws.sheet_view.rightToLeft = True

    headers = ["ردیف", "نام شرکت", "کشور", "نوع", "گرید پیشنهادی", "وب‌سایت", "ایمیل",
               "واتساپ/تلفن", "وضعیت تماس", "دلیل (مدل)", "دسته‌بندی‌های سایت",
               "توضیحات", "غرفه", "لینک پروفایل"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, c in enumerate(companies, start=1):
        ws.append([
            i,
            c.get("name", ""),
            c.get("country", ""),
            COMPANY_TYPE_FA.get(c.get("company_type"), c.get("company_type") or ""),
            c.get("suggested_grade", ""),
            c.get("website", ""),
            c.get("email", ""),
            c.get("whatsapp", ""),
            c.get("contact_note", "") if not (c.get("email") or c.get("whatsapp")) else "",
            c.get("llm_reason", ""),
            ", ".join(c.get("categories", [])[:6]),
            (c.get("description") or "")[:300],
            c.get("stand", ""),
            c.get("profile_url", ""),
        ])

    widths = [6, 32, 16, 16, 30, 30, 26, 20, 30, 40, 40, 55, 20, 45]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_path = os.path.join(EXTERNAL_XLSX_DIR, f"{safe_filename(event_name)}.xlsx")
    tmp_path = out_path + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, out_path)
    return out_path


def write_output_from_progress(out_path: str, event_name: str, source_url: str,
                                total_candidates: int, progress: dict) -> int:
    slug_prefix = re.sub(r"[^a-z0-9]+", "-", (event_name or "event").lower()).strip("-")
    companies = [v for v in progress.values() if v.get("relevant")]
    companies.sort(key=lambda c: c.get("_order", 0))
    companies = [
        {k: v for k, v in c.items() if k not in ("_order", "rule_based_grade", "relevant")}
        for c in companies
    ]
    for i, item in enumerate(companies, start=1):
        item["id"] = f"{slug_prefix}-{i}"
    output = {
        "event_name": event_name,
        "source_url": source_url,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "total_candidates_after_rules": total_candidates,
        "total_relevant": len(companies),
        "total_processed": len(progress),
        "companies": companies,
    }
    atomic_write_json(out_path, output)
    write_xlsx(companies, event_name)
    return len(companies)


def llm_refine(items: list[dict], progress: dict, progress_path: str,
               on_batch_done, generate_fn, deadline: float | None = None) -> None:
    """
    زیرمجموعه‌ی از قبل غربال‌شده را دسته‌دسته به مدل زبانی (generate_fn) می‌دهد تا
    واقعاً تولیدکننده/بازرگان بودن و گرید دقیق را تأیید کند. نتیجه مستقیم در
    progress (دیکشنری ورودی/خروجی، کلید = candidate_key) می‌نویسه و بعد از هر
    دسته فوری ذخیره می‌کنه — اگه اجرا وسط راه قطع بشه، هیچ کاری از دست نمی‌ره و
    اجرای بعدی از همین‌جا ادامه می‌ده.

    آیتم‌هایی که از قبل در progress هستن (چه تأییدشده چه ردشده) دوباره به مدل
    فرستاده نمی‌شن. آیتم‌هایی که هیچ توضیحی ندارن (فقط اسم/دسته) هم فرستاده
    نمی‌شن — مدل چیزی برای قضاوت اضافه بر خودِ برچسب دسته نداره؛ به‌جاش مستقیم
    با rule_based_grade خودشون در progress ثبت می‌شن.

    deadline: زمان یونیکس (time.time()) که اگه از آن گذشتیم، به‌جای کرش/قطع‌شدن
    ناگهانی (مثلاً با هارد‌تایم‌اوت CI)، تمیز و با آخرین checkpoint متوقف می‌شیم.
    """
    todo = [it for it in items if candidate_key(it) not in progress]
    print(f"  {len(items) - len(todo)} شرکت از اجرای قبلی در progress موجوده و رد می‌شه.")

    has_desc = [len((it.get("description") or "").strip()) >= 15 for it in todo]
    with_desc = [it for it, ok in zip(todo, has_desc) if ok]
    without_desc = [it for it, ok in zip(todo, has_desc) if not ok]
    print(f"  {len(with_desc)} شرکت جدید با توضیحات کافی به مدل داده می‌شه؛ "
          f"{len(without_desc)} شرکت بدون توضیح فقط با برچسب دسته ثبت می‌شه.")

    order_base = len(progress)
    for i, it in enumerate(without_desc):
        key = candidate_key(it)
        relevant = bool(it.get("rule_based_grade"))
        progress[key] = {
            **it, "relevant": relevant, "company_type": None,
            "suggested_grade": it.get("rule_based_grade"),
            "llm_reason": "بدون توضیح در پروفایل؛ فقط بر اساس دسته‌بندی سایت" if relevant else "بدون توضیح و بدون دسته‌ی مرتبط",
            "_order": order_base + i,
        }
    if without_desc:
        save_progress(progress_path, progress)
        on_batch_done()

    total_batches = (len(with_desc) + LLM_BATCH_SIZE - 1) // LLM_BATCH_SIZE
    order_base = len(progress)

    for b in range(total_batches):
        if deadline and time.time() > deadline:
            print(f"  [DEADLINE] زمان تمام شد؛ {total_batches - b} بچ باقی‌مانده برای اجرای بعدی می‌مونه.")
            break

        batch = with_desc[b * LLM_BATCH_SIZE: (b + 1) * LLM_BATCH_SIZE]
        payload = [
            {
                "index": i,
                "name": it["name"],
                "categories": it.get("categories", [])[:5],
                "description": (it.get("description") or "")[:160],
            }
            for i, it in enumerate(batch)
        ]

        try:
            raw = generate_fn(LLM_SYSTEM_PROMPT, json.dumps(payload, ensure_ascii=False))
            verdicts = extract_json_array(raw)
            by_index = {v.get("index"): v for v in verdicts if isinstance(v, dict)}
        except Exception as e:  # noqa: BLE001 — یک بچ خراب نباید کل اجرا را متوقف کند
            print(f"  [WARN] بچ {b + 1}/{total_batches} با خطا رد شد (بعداً دوباره تلاش می‌شه): {e}")
            continue

        kept = 0
        for i, it in enumerate(batch):
            verdict = by_index.get(i) or {}
            relevant = bool(verdict.get("relevant"))
            key = candidate_key(it)
            progress[key] = {
                **it,
                "relevant": relevant,
                "company_type": verdict.get("company_type"),
                "suggested_grade": verdict.get("grade") or it.get("rule_based_grade"),
                "llm_reason": verdict.get("reason") or ("بدون پاسخ مدل برای این مورد" if i not in by_index else None),
                "_order": order_base + b * LLM_BATCH_SIZE + i,
            }
            kept += relevant

        save_progress(progress_path, progress)
        total_so_far = on_batch_done()
        print(f"  [OK] بچ {b + 1}/{total_batches}: {kept}/{len(batch)} تأیید شد "
              f"(مجموع مرتبط تا الان: {total_so_far})")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("url", help="آدرس صفحه‌ی Exhibitor List نمایشگاه")
    parser.add_argument("--max", type=int, default=None, help="حداکثر تعداد شرکت برای دریافت (برای تست)")
    parser.add_argument("--event-name", default=None, help="نام نمایشگاه برای درج در خروجی (پیش‌فرض: اسلاگ URL)")
    parser.add_argument("--no-llm", action="store_true", help="پاس نهایی مدل زبانی را رد کن (فقط فیلتر قانون‌محور)")
    parser.add_argument("--engine", choices=["gemini", "groq"], default="gemini",
                         help="موتور پاس نهایی (پیش‌فرض Gemini — Groq سهمیه‌ی مشترک با بخش /ask زنده‌ی سایت داره)")
    parser.add_argument("--no-enrich", action="store_true",
                         help="استخراج ایمیل/وب‌سایت/واتساپ از سایت هر شرکت را رد کن")
    parser.add_argument("--deadline-minutes", type=float, default=None,
                         help="بعد از این‌قدر دقیقه، تمیز با آخرین checkpoint متوقف شو (برای اجرای زیر سقف زمانی CI)")
    args = parser.parse_args()

    deadline = (time.time() + args.deadline_minutes * 60) if args.deadline_minutes else None

    load_env()
    base_url, event_slug = parse_event_slug(args.url)
    event_name = args.event_name or event_slug

    print(f"[1/5] گرفتن درخت دسته‌بندی از {args.url}")
    with httpx.Client(follow_redirects=True) as client:
        categories, _ = fetch_category_tree(client, base_url, event_slug, args.url)
        top_ids, sub_ids = select_relevant_categories(categories)
        print(f"  {len(categories)} دسته/زیردسته پیدا شد؛ {len(sub_ids)} زیردسته‌ی مشخص انتخاب شد "
              f"(از {len(LEAF_CATEGORY_TO_GRADE)} مورد در فهرست).")
        if not sub_ids:
            print("  [WARN] هیچ‌کدام از زیردسته‌های فهرست LEAF_CATEGORY_TO_GRADE روی این سایت پیدا نشد؛ "
                  "شاید تاکسونومی این نمایشگاه با Gulfood فرق دارد و باید فهرست را برایش تنظیم کرد.")

        print("[2/5] دریافت شرکت‌ها (فیلترشده روی زیردسته‌های مرتبط، سمت سرور)")
        raw_items = fetch_all_exhibitors(
            client, base_url, event_slug, args.url,
            new_category=",".join(top_ids), new_sub_category=",".join(sub_ids),
            max_items=args.max,
        )

    print(f"[3/5] حذف موارد تکراری از {len(raw_items)} مورد خام")
    unique_items = dedupe_by_profile(raw_items)
    print(f"  {len(unique_items)} شرکت یکتا")

    candidates = tag_rule_based_grade(unique_items)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, f"{event_slug}.json")
    progress_path = os.path.join(OUTPUT_DIR, f"{event_slug}.progress.json")
    progress = load_progress(progress_path)
    if progress:
        print(f"  [RESUME] {len(progress)} شرکت از اجرای قبلی در {progress_path} پیدا شد.")

    def flush_output() -> int:
        return write_output_from_progress(out_path, event_name, args.url, len(candidates), progress)

    if args.no_llm:
        print("[4/5] رد شد (--no-llm) — خروجی فقط بر اساس قانون است")
        order_base = len(progress)
        for i, it in enumerate(candidates):
            key = candidate_key(it)
            if key in progress:
                continue
            progress[key] = {
                **it, "relevant": bool(it["rule_based_grade"]),
                "suggested_grade": it["rule_based_grade"], "llm_reason": None,
                "_order": order_base + i,
            }
        save_progress(progress_path, progress)
    else:
        print(f"[4/5] تأیید نهایی با {args.engine} روی {len(candidates)} شرکت (دسته‌ای، هر دسته {LLM_BATCH_SIZE} تا)")
        generate_fn = make_generate_fn(args.engine)
        llm_refine(candidates, progress, progress_path, on_batch_done=flush_output,
                   generate_fn=generate_fn, deadline=deadline)

    total_relevant = flush_output()
    print(f"  {total_relevant} شرکت مرتبط تا الان (از {len(progress)} پردازش‌شده از {len(candidates)})")

    if args.no_enrich:
        print("[5/5] رد شد (--no-enrich)")
    else:
        print("[5/5] استخراج ایمیل/وب‌سایت/واتساپ برای شرکت‌های مرتبط")
        enrich_contacts_pass(progress, progress_path, on_batch_done=flush_output, deadline=deadline)

    total_relevant = flush_output()
    print(f"\n[DONE] {total_relevant} شرکت مرتبط در {out_path} ذخیره شد "
          f"(مجموع پردازش‌شده تا الان: {len(progress)} از {len(candidates)}).")


if __name__ == "__main__":
    main()
