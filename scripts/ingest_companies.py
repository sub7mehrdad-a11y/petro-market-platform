"""
یکدست‌سازی فایل‌های اکسل «لیست شرکت‌ها» (که ستون‌هاشون بین کشورها کمی فرق داره)
به یک اسکیمای مشترک در data/companies.json.

چرا اینجا از هوش مصنوعی استفاده نشده: چون داده از قبل جدولیه (اکسل تمیز)،
فقط اسم ستون‌ها فرق می‌کنه — یک نگاشت دستی ساده کافی و قابل‌اعتمادتر از AI هست.

دو نوع منبع داریم:
  1. SOURCES — یک فایل = یک کشور (مثل قبل). برای اضافه‌کردن کشور جدید یک ردیف
     اضافه کن.
  2. GLOBAL_SOURCES — یک فایل = چند کشور با هم (مثلاً فهرست‌های Volza/Trade Map
     که خروجی یک جست‌وجوی جهانی‌ان، نه مخصوص یک بازار). این‌جا کشور از خودِ
     ستون آدرس هر ردیف («…، نام‌کشور») با نگاشت country_name_map.json استخراج
     می‌شه — همون فایلی که ingest_trade_map.py هم استفاده می‌کنه، تا اسم فارسی
     کشورها بین بخش‌های مختلف سایت یکسان بمونه.

اگه ستون‌های اکسل جدید اسم متفاوتی داشتن، به FIELD_KEYWORDS اضافه کن.
"""

import os
import re
import json

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
REPORTS_SRC_DIR = os.path.join(BASE_DIR, "گزارش", "لیست شرکتها")
NAME_MAP_FILE = os.path.join(os.path.dirname(__file__), "country_name_map.json")
OUTPUT_FILE = os.path.join(BASE_DIR, "data", "companies.json")

SOURCES = [
    {"file": "Brazil_Companies.xlsx", "country": "برزیل"},
    {"file": "Jordan_Companies.xlsx", "country": "اردن"},
    {"file": "kenya_Companies.xlsx", "country": "کنیا"},
    {"file": "شرکت های عراقی.xlsx", "country": "عراق"},
    {"file": "Syria_companies.xlsx", "country": "سوریه"},
    {"file": "Armenia_Companies_1.xlsx", "country": "ارمنستان"},
]

# فایل‌های جهانی — کشور هر شرکت از ستون آدرس همان ردیف استخراج می‌شود، نه از
# اسم فایل. منبع هر فایل را همین‌جا یادداشت کن تا بعداً معلوم باشد از کجا آمده.
GLOBAL_SOURCES = [
    {"file": "Verified_Global_Companies.xlsx", "source_note": "Volza (راستی‌آزمایی‌شده)"},
    {"file": "TradeMap Companies Database - Verified & Completed.xlsx", "source_note": "ITC Trade Map (راستی‌آزمایی‌شده)"},
]

# چون سرستون‌ها بین فایل‌ها فرق می‌کنن (مثلاً "آدرس" در فارسی یا "Address" در
# انگلیسی، یا با توضیح فارسی داخل پرانتز)، با کلیدواژه تشخیص می‌دیم نه تطابق دقیق.
FIELD_KEYWORDS = {
    "english_name": ["english name"],
    "arabic_name": ["arabic name"],
    "industry": ["industry", "صنعت"],
    "target_grade": ["target grade", "گرید هدف"],
    "purchasing_potential": ["purchasing potential", "پتانسیل خرید"],
    "action_plan": ["action plan", "برنامه اقدام"],
    "address": ["address", "آدرس"],
    "phone": ["phone", "شماره تماس", "تلفن"],
    "email": ["email", "ایمیل"],
    "website": ["website", "وب‌سایت", "وبسایت"],
}

# نام‌های رایج/کوتاه کشورها که در country_name_map.json (که نام رسمی ISO دارد،
# مثل "Viet Nam" یا "Korea, Republic of") مستقیم پیدا نمی‌شوند.
COUNTRY_ALIASES = {
    "vietnam": "Viet Nam",
    "uae": "United Arab Emirates",
    "south korea": "Korea, Republic of",
    "north korea": "Korea, Democratic People's Republic of",
    "russia": "Russian Federation",
    "iran": "Iran, Islamic Republic of",
    "syria": "Syrian Arab Republic",
    "laos": "Lao People's Democratic Republic",
    "moldova": "Moldova, Republic of",
    "tanzania": "Tanzania, United Republic of",
    "bolivia": "Bolivia, Plurinational State of",
    "venezuela": "Venezuela, Bolivarian Republic of",
    "brunei": "Brunei Darussalam",
    "usa": "United States of America",
    "u.s.a.": "United States of America",
    "uk": "United Kingdom",
}

# پسوندهایی که در نام رسمی کشور بعد از ویرگول می‌آیند (مثل "Korea, Republic of")
# — اگر آخرین بخشِ آدرس فقط همین پسوند باشد (چون آدرس هم با ویرگول جدا شده)،
# باید با بخش قبلی ترکیب بشه، وگرنه به‌تنهایی هیچ کشوری نیست.
REPUBLIC_SUFFIXES = {
    "republic of", "democratic republic of", "islamic republic of",
    "united republic of", "bolivarian republic of", "plurinational state of",
}


def map_headers(header_row):
    mapping = {}  # column index -> field name
    for idx, cell in enumerate(header_row):
        if not cell:
            continue
        cell_lower = str(cell).lower()
        for field, keywords in FIELD_KEYWORDS.items():
            if any(kw in cell_lower for kw in keywords):
                mapping[idx] = field
                break
    return mapping


def clean_value(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return str(v).strip()


def _clean_address_segment(seg):
    """پرانتز انتهایی («(unverified)») و کد پستی چسبیده («Singapore 609933») را حذف می‌کند."""
    seg = seg.strip()
    seg = re.sub(r"\s*\(.*?\)\s*$", "", seg)
    seg = re.sub(r"\s+\d{3,}.*$", "", seg)
    return seg.strip()


def resolve_country_from_address(address, name_map):
    """
    از ستون آدرس (که معمولاً با «…، شهر، کشور» تمام می‌شه) اسم فارسی/ISO2 کشور
    رو استخراج می‌کنه. از انتهای آدرس به سمت جلو می‌گرده تا اولین بخشی که یک
    کشور شناخته‌شده باشه رو پیدا کنه — چون بعضی آدرس‌ها با کد پستی یا جزئیات
    شعبه تموم می‌شن (مثلاً «...، Canada, H9X 3P1»).
    """
    if not address:
        return None
    parts = [p.strip() for p in address.split(",") if p.strip()]
    if not parts:
        return None

    if parts[-1].lower() in REPUBLIC_SUFFIXES and len(parts) >= 2:
        parts = parts[:-2] + [f"{parts[-2]}, {parts[-1]}"]

    for raw in reversed(parts):
        seg = _clean_address_segment(raw)
        if not seg:
            continue
        if seg in name_map:
            return name_map[seg]
        alias = COUNTRY_ALIASES.get(seg.lower())
        if alias and alias in name_map:
            return name_map[alias]
    return None


def load_companies(path, country):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_map = map_headers(rows[0])
    companies = []
    for row in rows[1:]:
        if not any(row):
            continue
        record = {"country": country}
        for idx, field in header_map.items():
            if idx < len(row):
                record[field] = clean_value(row[idx])
        if not record.get("english_name"):
            continue
        companies.append(record)
    return companies


def load_global_companies(path, source_note, name_map):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    header_map = map_headers(rows[0])
    companies = []
    unresolved = []
    seen_exact = set()

    for row in rows[1:]:
        if not any(row):
            continue

        # ردیف‌های کاملاً تکراری (همه‌ی ستون‌ها یکسان) را رد کن — تفاوت داشتن
        # در حتی یک ستون (مثلاً شعبه‌ی دیگر همان شرکت با آدرس/تلفن متفاوت) کافیه
        # که نگهش داریم، چون برای بازاریابی نقطه‌تماس واقعاً متفاوتیه.
        dedup_key = tuple((str(v).strip().lower() if v is not None else "") for v in row)
        if dedup_key in seen_exact:
            continue
        seen_exact.add(dedup_key)

        record = {"source_note": source_note}
        for idx, field in header_map.items():
            if idx < len(row):
                record[field] = clean_value(row[idx])
        if not record.get("english_name"):
            continue

        country_info = resolve_country_from_address(record.get("address"), name_map)
        if not country_info:
            unresolved.append((record["english_name"], record.get("address")))
            continue

        record["country"] = country_info["fa"]
        companies.append(record)

    return companies, unresolved


def main():
    with open(NAME_MAP_FILE, "r", encoding="utf-8") as f:
        name_map = json.load(f)

    all_companies = []

    for src in SOURCES:
        path = os.path.join(REPORTS_SRC_DIR, src["file"])
        if not os.path.exists(path):
            print(f"[WARN] فایل پیدا نشد: {path}")
            continue
        companies = load_companies(path, src["country"])
        print(f"[OK] {src['country']}: {len(companies)} شرکت")
        all_companies.extend(companies)

    for src in GLOBAL_SOURCES:
        path = os.path.join(REPORTS_SRC_DIR, src["file"])
        if not os.path.exists(path):
            print(f"[WARN] فایل پیدا نشد: {path}")
            continue
        companies, unresolved = load_global_companies(path, src["source_note"], name_map)
        print(f"[OK] {src['file']} ({src['source_note']}): {len(companies)} شرکت در "
              f"{len({c['country'] for c in companies})} کشور")
        if unresolved:
            print(f"  [WARN] {len(unresolved)} ردیف بدون کشور قابل‌تشخیص (رد شدند):")
            for name, addr in unresolved:
                print(f"    - {name}: {addr}")
        all_companies.extend(companies)

    for i, c in enumerate(all_companies, start=1):
        c["id"] = f"company-{i}"

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(all_companies, f, ensure_ascii=False, indent=2)

    print(f"\n[DONE] مجموعاً {len(all_companies)} شرکت در {OUTPUT_FILE} ذخیره شد.")


if __name__ == "__main__":
    main()
