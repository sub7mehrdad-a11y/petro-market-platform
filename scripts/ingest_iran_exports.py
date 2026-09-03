"""
ETL آمار رسمی گمرک جمهوری اسلامی ایران (IRICA) برای صادرات جوش شیرین ایران
(کد تعرفه ۲۸۳۶۳۰۹۰: «سایر هیدروژن کربنات سدیم، بجز بی‌کربنات غیرتزریقی گرید دارویی»
— یعنی همه‌ی گریدها به‌جز گرید دارویی تزریقی؛ کاربر این رو «گرید خوراکی» می‌نامه).

چرا دو نوع فایل با گرانولاریتی متفاوت داریم:
  - export-detail-1404-10m.xlsx: ریزترین سطح (سال، گمرک، کشور، تعرفه، وزن، ارزش
    ریالی/دلاری) برای ده‌ماهه‌ی ۱۴۰۴ — تنها فایلی که تفکیک کشور دار.
  - export-by-tariff-140X.xlsx: فقط جمع کل کشوری به‌ازای هر تعرفه، برای سال کامل
    ۱۴۰۲ و ۱۴۰۳ — تفکیک کشور نداره، ولی سال کامله (نه ده‌ماهه).
  - export-by-country-1404-10m-partial.xlsx: زیرمجموعه‌ی همون export-detail (بدون
    ارزش دلاری) که کاربر اول فرستاد — عمداً اینجا استفاده نمی‌شه چون export-detail
    همون داده رو کامل‌تر داره (با قیمت). فقط برای مرجع/شفافیت نگه داشته شده.
  - export-by-customs-office-1404-10m.xlsx: همون بازه‌ی ده‌ماهه‌ی ۱۴۰۴ (مجموع کلش
    دقیقاً با export-detail یکسانه — تأیید متقاطع)، ولی از قبل به تفکیک گمرک
    صادرکننده + گرید (اصلی/دارویی) آماده شده؛ اینجا فقط برای بخش «گمرک صادرکننده»
    و رقم مکمل «گرید دارویی» استفاده می‌شه، نه برای تفکیک کشور (که از export-detail
    میاد و قبلاً اعتبارسنجی شده).

اعداد این اسکریپت با ارقام رسمی‌ای که کاربر مستقیماً از گمرک نقل کرد اعتبارسنجی شد:
  ۱۴۰۲ = ۲۳٬۷۰۰ تن، ۱۴۰۳ = ۳۵٬۵۰۰ تن (خروجی اینجا: ۲۳٬۶۶۷ و ۳۵٬۴۹۴ — تطابق کامل).
"""

import os
import json

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
SRC_DIR = os.path.join(BASE_DIR, "گزارش", "گمرک ایران")
OUTPUT_FILE = os.path.join(BASE_DIR, "data", "iran_exports.json")

TARGET_HS = "28363090"
HS_DESCRIPTION_FA = "سایر هیدروژن کربنات سدیم (بی‌کربنات سدیم)، بجز بی‌کربنات غیرتزریقی گرید دارویی"

DETAIL_FILE = os.path.join(SRC_DIR, "export-detail-1404-10m.xlsx")
CUSTOMS_OFFICE_FILE = os.path.join(SRC_DIR, "export-by-customs-office-1404-10m.xlsx")
PHARMA_HS = "28363010"  # گرید دارویی (غیر تزریقی) — عمداً از آمار اصلی جدا نگه داشته می‌شه
TARIFF_FILES = [
    (os.path.join(SRC_DIR, "export-by-tariff-1402.xlsx"), "1402", True),
    (os.path.join(SRC_DIR, "export-by-tariff-1403.xlsx"), "1403", True),
]

# نام‌های رسمی گمرک (که با نام‌های استفاده‌شده در بقیه‌ی سایت فرق دارن، مثل
# «فدراسیون روسيه» به‌جای «روسیه») به نام فارسی + iso2 یکسان با بقیه‌ی پلتفرم.
# مقادیر عیناً از scripts/country_name_map.json (منبع مشترک بقیه‌ی ingest ها) گرفته شده.
CUSTOMS_NAME_MAP = {
    "فدراسیون روسيه": {"fa": "روسیه", "iso2": "ru"},
    "عراق": {"fa": "عراق", "iso2": "iq"},
    "افغانستان": {"fa": "افغانستان", "iso2": "af"},
    "ازبکستان": {"fa": "ازبکستان", "iso2": "uz"},
    "جمهوري آذربايجان": {"fa": "آذربایجان", "iso2": "az"},
    "بلاروس": {"fa": "بلاروس", "iso2": "by"},
    "امارات متحده عربي": {"fa": "امارات", "iso2": "ae"},
    "اوکراين": {"fa": "اوکراین", "iso2": "ua"},
    "ارمنستان": {"fa": "ارمنستان", "iso2": "am"},
    "گرجستان": {"fa": "گرجستان", "iso2": "ge"},
    "تاجيکستان": {"fa": "تاجیکستان", "iso2": "tj"},
    "کرواسي (Hrvatska)": {"fa": "کرواسی", "iso2": "hr"},
    "ترکمنستان": {"fa": "ترکمنستان", "iso2": "tm"},
    "ترکيه": {"fa": "ترکیه", "iso2": "tr"},
    "اردن": {"fa": "اردن", "iso2": "jo"},
    "قرقيزستان": {"fa": "قرقیزستان", "iso2": "kg"},
    "يمن": {"fa": "یمن", "iso2": "ye"},
    "ایتالیا": {"fa": "ایتالیا", "iso2": "it"},
    "نيجريه": {"fa": "نیجریه", "iso2": "ng"},
    "آلمان": {"fa": "آلمان", "iso2": "de"},
}


def parse_detail_by_country(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["آمار"]
    totals: dict[str, list] = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or len(row) < 8:
            continue
        _year, _customs, country_raw, tariff, _desc, kg, rial, usd = row
        if tariff != TARGET_HS or not country_raw:
            continue
        bucket = totals.setdefault(country_raw, [0.0, 0.0, 0.0])
        bucket[0] += kg or 0
        bucket[1] += rial or 0
        bucket[2] += usd or 0

    rows = []
    unmapped = []
    for country_raw, (kg, rial, usd) in totals.items():
        info = CUSTOMS_NAME_MAP.get(country_raw)
        if not info:
            unmapped.append(country_raw)
            continue
        tons = round(kg / 1000, 1)
        rows.append({
            "country_fa": info["fa"],
            "iso2": info["iso2"],
            "tons": tons,
            "value_usd": round(usd),
            "unit_price_usd_per_ton": round(usd / tons, 1) if tons else None,
        })

    if unmapped:
        print(f"[WARN] {len(unmapped)} کشور بدون نگاشت (رد شدند): {unmapped}")

    rows.sort(key=lambda r: -r["tons"])
    return rows


def parse_by_customs_office(path: str) -> tuple[list[dict], dict]:
    """
    گمرک صادرکننده‌ی هر محموله رو تجمیع می‌کنه — سؤال «کدوم مرز/بندر ایران
    بیشترین صادرات جوش شیرین رو داره»، که هیچ‌جای دیگه‌ی سایت جواب داده نمی‌شه.

    فقط HS اصلی (TARGET_HS) رو حساب می‌کنه تا با بقیه‌ی آمار سایت (که همون
    محدوده رو داره) هم‌راستا بمونه؛ گرید دارویی (PHARMA_HS) جدا و به‌عنوان یک
    رقم مکمل کوچیک برمی‌گرده، نه قاطی‌شده با تناژ اصلی.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    offices: dict[str, list] = {}
    pharma = [0.0, 0.0]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or row[0] is None:
            continue
        office, _country, tariff, _grade, _desc, kg, _tons, usd, _rial, _rate, _share = row
        tariff = str(int(tariff)) if tariff is not None else ""
        if tariff == PHARMA_HS:
            pharma[0] += (kg or 0) / 1000
            pharma[1] += usd or 0
            continue
        if tariff != TARGET_HS:
            continue
        bucket = offices.setdefault(office, [0.0, 0.0])
        bucket[0] += (kg or 0) / 1000
        bucket[1] += usd or 0

    result = [
        {"office_fa": office, "tons": round(tons, 1), "value_usd": round(usd)}
        for office, (tons, usd) in offices.items()
    ]
    result.sort(key=lambda r: -r["tons"])

    pharma_summary = {"tons": round(pharma[0], 1), "value_usd": round(pharma[1])}
    return result, pharma_summary


def parse_tariff_total(path: str) -> dict | None:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row and row[0] == TARGET_HS:
            _tariff, _desc, kg, rial, usd = row
            return {"tons": round(kg / 1000, 1), "value_rial": round(rial), "value_usd": round(usd)}
    return None


def main():
    destinations = parse_detail_by_country(DETAIL_FILE)
    total_1404_tons = round(sum(r["tons"] for r in destinations), 1)
    total_1404_usd = round(sum(r["value_usd"] for r in destinations))

    annual_totals = []
    for path, year_fa, is_partial in TARIFF_FILES:
        result = parse_tariff_total(path)
        if not result:
            print(f"[WARN] تعرفه {TARGET_HS} توی {path} پیدا نشد.")
            continue
        annual_totals.append({
            "year_fa": year_fa,
            "is_partial_year": is_partial,
            **result,
        })
        print(f"[OK] {year_fa}: {result['tons']:,.1f} تن، {result['value_usd']:,.0f} دلار")

    annual_totals.append({
        "year_fa": "1404",
        "is_partial_year": True,
        "months_covered": 10,
        "tons": total_1404_tons,
        "value_usd": total_1404_usd,
        "value_rial": None,
    })
    print(f"[OK] 1404 (10 ماهه): {total_1404_tons:,.1f} تن، {total_1404_usd:,.0f} دلار")

    customs_offices, pharma_summary = parse_by_customs_office(CUSTOMS_OFFICE_FILE)
    print(f"[OK] {len(customs_offices)} گمرک صادرکننده؛ گرید دارویی: "
          f"{pharma_summary['tons']:,.1f} تن، {pharma_summary['value_usd']:,.0f} دلار")

    output = {
        "hs_code": TARGET_HS,
        "hs_description_fa": HS_DESCRIPTION_FA,
        "source": "گمرک جمهوری اسلامی ایران (irica.ir)",
        "annual_totals": annual_totals,
        "destinations_1404_10m": destinations,
        "customs_offices_1404_10m": customs_offices,
        "pharma_grade_1404_10m": pharma_summary,
    }

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n[DONE] {len(destinations)} کشور مقصد، {OUTPUT_FILE} ذخیره شد.")


if __name__ == "__main__":
    main()
